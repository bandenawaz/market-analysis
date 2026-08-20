"""
Excel export functionality for Satta Matka data
Enhanced to include live market data alongside bet data
"""
import io
import datetime
from typing import Dict, Any, List
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYLOX_AVAILABLE = True
except ImportError:
    OPENPYLOX_AVAILABLE = False
    logging.warning("openpyxl not available. Excel export functionality will be limited.")

from market_fetcher import get_market_data

logger = logging.getLogger(__name__)

class ExcelExporter:
    def __init__(self):
        if not OPENPYLOX_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export functionality")
    
    def create_bet_data_sheet(self, workbook, bet_data: List[Dict[str, Any]]):
        """Create a worksheet for bet tracking data"""
        sheet = workbook.create_sheet(title="Bet Tracking")
        
        # Headers
        headers = ['Timestamp', 'User', 'Market', 'Bet Type', 'Amount', 'Status', 'Payout']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, bet in enumerate(bet_data, 2):
            sheet.cell(row=row, column=1, value=bet.get('timestamp', ''))
            sheet.cell(row=row, column=2, value=bet.get('user', ''))
            sheet.cell(row=row, column=3, value=bet.get('market', ''))
            sheet.cell(row=row, column=4, value=bet.get('bet_type', ''))
            sheet.cell(row=row, column=5, value=bet.get('amount', 0))
            sheet.cell(row=row, column=6, value=bet.get('status', ''))
            sheet.cell(row=row, column=7, value=bet.get('payout', 0))
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        return sheet
    
    def create_market_data_sheet(self, workbook):
        """Create a worksheet for live market data"""
        sheet = workbook.create_sheet(title="Live Market Data")
        
        # Get current market data
        market_data = get_market_data()
        
        # Headers
        headers = ['Market', 'Open', 'Jodi', 'Close', 'Final Ank', 'Last Updated']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        row = 2
        for market_name, data in market_data.items():
            sheet.cell(row=row, column=1, value=market_name)
            sheet.cell(row=row, column=2, value=data.get('open', ''))
            sheet.cell(row=row, column=3, value=data.get('jodi', ''))
            sheet.cell(row=row, column=4, value=data.get('close', ''))
            sheet.cell(row=row, column=5, value=data.get('final_ank', ''))
            
            # Format timestamp
            timestamp = data.get('timestamp', 0)
            if timestamp:
                dt = datetime.datetime.fromtimestamp(timestamp)
                sheet.cell(row=row, column=6, value=dt.strftime('%Y-%m-%d %H:%M:%S'))
            else:
                sheet.cell(row=row, column=6, value='')
            
            row += 1
        
        # Add timestamp for when data was fetched
        sheet.cell(row=row+2, column=1, value="Data fetched at:")
        sheet.cell(row=row+2, column=2, value=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        return sheet
    
    def export_to_excel(self, bet_data: List[Dict[str, Any]] = None) -> bytes:
        """
        Export bet data and market data to Excel format
        Returns Excel file as bytes
        """
        if bet_data is None:
            bet_data = []  # Empty bet data if none provided
        
        # Create workbook
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create bet data sheet
        if bet_data:
            self.create_bet_data_sheet(workbook, bet_data)
        else:
            # Create empty bet tracking sheet with headers
            sheet = workbook.create_sheet(title="Bet Tracking")
            headers = ['Timestamp', 'User', 'Market', 'Bet Type', 'Amount', 'Status', 'Payout']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
        
        # Create market data sheet
        self.create_market_data_sheet(workbook)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()

# Convenience function for external use
def export_bet_and_market_data(bet_data: List[Dict[str, Any]] = None) -> bytes:
    """Export bet data and market data to Excel format"""
    exporter = ExcelExporter()
    return exporter.export_to_excel(bet_data)
